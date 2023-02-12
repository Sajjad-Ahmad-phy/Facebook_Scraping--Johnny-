import os
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import pyautogui
from time import sleep

option = webdriver.ChromeOptions()
option.add_argument('--disable-blink-features=AutomationControlled')
option.add_argument('--ignore-certificate-errors')
option.add_argument('--user-data-dir=C:\\Users\\lenovo\\AppData\\Local\\Google\\Chrome\\User Data\\Default')
option.add_argument('--profile-directory=Person 1')
option.add_argument('--disable-blink-features=AutomationControlled')
driver = webdriver.Chrome()

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

"Requesting Facebook"
driver.get("https://www.facebook.com/")
driver.maximize_window()

def Login_Facebook(username, password):
    driver.find_element('xpath', """ //*[@id="email"]""").send_keys(username)
    driver.find_element('xpath', """  //*[@id="pass"]""").send_keys(password)
    driver.find_element(By.NAME, """login""").click()
    sleep(2)


"**************Data Scraping*****************************"
def OutPut_Data(ACT_Sheet,sheet_N, Name_of_Output_Excel_file, W_B):

    try:
        Col = sheet_N['A']
        Col_list = [Col[x].value for x in range(len(Col))]

        for url in range(len(ACT_Sheet['A'])):
            if url not in Col_list:
                row_no = url+2
                c = ACT_Sheet.cell(row=row_no, column=1)
                URL = f"{c.value}"
                driver.get(URL)
                sleep(1)
                Final_List = []
                Final_List.append(URL)

                data = driver.find_element(By.XPATH, " /html/body ")
                List = data.text.split('\n')
                sleep(1)

                if List[32] == 'This listing is no longer available':
                    continue

                try:
                    "Appending Title price and listed data"
                    Final_List.append(List[31])
                    Final_List.append(List[32])
                    Final_List.append(List[33])
                except:
                    pass
                try:
                    D1 = List.index("Seller's description")
                    D2 = List.index("Location is approximate")
                    Des1 = List[D1+1:D2-1]
                    "Seller's description"
                    Description = " ".join([str(elem) for elem in Des1])
                    Final_List.append(Description)
                except:
                    Final_List.append('This listing is no longer available')

                try:
                    "Engine details ..."
                    a1 = List.index('About this vehicle')
                    a2 = List.index("Seller's description")
                    about = List[a1:a2 - 1]
                    About_the_Vehicle = " ".join([str(ele) for ele in about])
                    Final_List.append(About_the_Vehicle)

                except:
                    pass
                try:
                    # links = driver.find_elements(By.TAG_NAME, 'img')
                    links = driver.find_elements(By.TAG_NAME, 'img')
                    for link in links:
                        href = link.get_attribute('src')
                        if href is not None:
                            Final_List.append(href)
                except:
                    pass

                sheet_N.append(Final_List)
                clear_screen()
                print(url)

                "Save o/p Excel"
                W_B.save(f"D:\\PYTHON PROJECTS\\Facebook_Scraping--Johnny-\\Excel files\\{Name_of_Output_Excel_file}.xlsx")

    except:
        pass


"**************Links scraping**********************"

class Links_Scraping_Functions:

    def __init__(self, CarModel_Url):
        self.Carmodel_Url = CarModel_Url

    def Output_Link(self):
        driver.get(self.Carmodel_Url)
        try:
            # while True:
            #     driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
            last_height = driver.execute_script("return document.body.scrollHeight")
            while True:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

                sleep(2)
                # calculate the new scroll height and compare it with the previous scroll height
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
        except:
            pass


        self.List_of_links = []
        self.links = driver.find_elements(By.TAG_NAME, 'a')
        for link in self.links:
            href = link.get_attribute('href')
            if href is not None:
                 if f'{href[0:42]}' == 'https://www.facebook.com/marketplace/item/':
                    self.List_of_links.append(href)

    def Old_output_Exfile(self, Excel_file_name):
        self.Excel_file_name = Excel_file_name
        Old_WB = load_workbook(filename=f'D:\\PYTHON PROJECTS\\Facebook_Scraping--Johnny-\\Excel files\\{self.Excel_file_name}.xlsx')
        Old_Sheet = Old_WB.active
        for link in self.List_of_links:
            Old_Sheet.append([f'{link}'])

        Old_WB.save(f'D:\\PYTHON PROJECTS\\Facebook_Scraping--Johnny-\\Excel files\\{self.Excel_file_name}.xlsx')

    def New_output_Exfile(self, New_Excel_file_name):
        self.New_Excel_file_name = New_Excel_file_name
        New_WB = Workbook()
        New_Sheet = New_WB.active

        for link in self.List_of_links:
            New_Sheet.append([f'{link}'])

        New_WB.save(f'D:\\PYTHON PROJECTS\\Facebook_Scraping--Johnny-\\Excel files\\{self.New_Excel_file_name}.xlsx')
